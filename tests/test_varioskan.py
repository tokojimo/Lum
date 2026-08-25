from io import BytesIO

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from luxplate.varioskan import (
    assign_biological_pair_ids,
    combine_kinetic_tables,
    combine_time_point_tables,
    inspect_workbook,
    normalize_strain_name,
    parse_kinetic_workbook,
    parse_single_time_workbook,
    parse_time_file_name,
    organize_time_files,
    suggest_biological_pair_id,
)


@pytest.mark.parametrize(
    ("name", "expected"),
    [("260626_SCFM2Po_Rep1_t0.xlsx", ("260626_SCFM2Po_Rep1", 0)),
     ("run_t10_valid.xlsx", ("run", 10)),
     ("experiment_t0(3).xlsx", ("experiment", 0))],
)
def test_time_point_is_extracted_independently_of_optional_suffix(name, expected):
    assert parse_time_file_name(name) == expected


def test_time_files_are_grouped_by_experiment_and_sorted_numerically():
    groups, missing = organize_time_files([
        "rep1_t10.xlsx", "rep2_t1.xlsx", "rep1_t2.xlsx", "rep1_t1.xlsx",
    ])
    assert [index for index, _ in groups["rep1"]] == [1, 2, 10]
    assert [index for index, _ in groups["rep2"]] == [1]
    assert missing["rep1"] == list(range(3, 10))


def test_time_file_validation_resolves_validated_version_and_rejects_ambiguous_duplicate():
    with pytest.raises(ValueError, match="ne contient pas"):
        organize_time_files(["rep1.xlsx"])
    groups, _ = organize_time_files(["rep1_t2.xlsx", "rep1_t2_valid.xlsx"])
    assert groups["rep1"] == [(2, "rep1_t2_valid.xlsx")]
    assert groups.duplicate_resolutions[("rep1", 2)] == (
        "rep1_t2_valid.xlsx", ("rep1_t2.xlsx", "rep1_t2_valid.xlsx"),
    )
    with pytest.raises(ValueError, match="Deux fichiers.*t0"):
        organize_time_files(["experiment_t0(1).xlsx", "experiment_t0(3).xlsx"])


def test_time_groups_are_stable_when_shuffled_and_keep_replicates_separate():
    ordered = [
        f"260702_SCFM2Po_Rep{rep}_t{index}.xlsx"
        for rep in range(1, 4) for index in range(8)
    ]
    shuffled = ordered[::2][::-1] + ordered[1::2]
    groups, missing = organize_time_files(shuffled)
    assert set(groups) == {f"260702_SCFM2Po_Rep{rep}" for rep in range(1, 4)}
    assert all([index for index, _ in files] == list(range(8)) for files in groups.values())
    assert all(points == [] for points in missing.values())


def test_validated_duplicate_does_not_manufacture_a_later_time_point():
    names = [f"260626_SCFM2Po_Rep1_t{index}.xlsx" for index in range(8)]
    names.insert(3, "260626_SCFM2Po_Rep1_t2_valid.xlsx")
    groups, _ = organize_time_files(names)
    assert [index for index, _ in groups["260626_SCFM2Po_Rep1"]] == list(range(8))
    assert groups["260626_SCFM2Po_Rep1"][2][1].endswith("_t2_valid.xlsx")


def test_biological_pair_suggestions_use_names_not_file_order():
    names = (
        "260616_SCFM2_screening_rep1", "260616_SCFM2_screening_rep2",
        "260618_SCFM2_screening_rep3", "260616_SCFM2kpi-80DMEM",
        "260616_SCFM2kpi-80DMEM_Rep2", "260616_SCFM2kpi-80DMEM_Rep3",
    )
    assert [suggest_biological_pair_id(name) for name in names] == [
        "bio1", "bio2", "bio3", "bio1", "bio2", "bio3",
    ]


def test_validated_biological_pair_mapping_is_attached_to_every_file_row():
    data = pd.DataFrame({"experience": ["run_rep1", "run_rep1", "other_Rep2"]})
    mapping = pd.DataFrame({
        "experience": ["run_rep1", "other_Rep2"],
        "biological_pair_id": ["donor-a", "donor-b"],
    })
    result = assign_biological_pair_ids(data, mapping)
    assert result["biological_pair_id"].tolist() == ["donor-a", "donor-a", "donor-b"]


def synthetic_workbook(*, second_luminescence=False) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    headers = ["Lecture en cours", "temps moy. [s]", "PAO1 (A01)", "PAO1 (A02)", "Blanc (B01)"]
    for name, scale in [("Absorbance 600", 0.01), ("Luminescence 1", 100.0)]:
        sheet = workbook.create_sheet(name)
        sheet.append(["export synthétique"])
        sheet.append(headers)
        sheet.append([1, 0, scale, scale * 2, scale / 2])
        sheet.append([2, 3600, scale * 3, scale * 4, scale / 2])
    if second_luminescence:
        sheet = workbook.create_sheet("Luminescence 2")
        sheet.append(headers)
        sheet.append([1, 0, 10, 20, 5])
        sheet.append([2, 3600, 30, 40, 5])
    plan = workbook.create_sheet("Plan de plaque")
    plan.append([])
    plan.append([])
    plan.append([])
    plan.append([None, *range(1, 13)])
    plan.append(["A", "PAO1", "PAO1"])
    plan.append([None, "Groupe 1", "Groupe 1"])
    plan.append(["B", "Blanc"])
    plan.append([None, "Groupe 1"])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def single_time_workbook(*, mismatched_name=False, missing_lum=False) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    names = ["Blanc1", "Blanc1", "Blanc1", *([None] * 9)]
    for sheet_name, marker, values in [
        ("Absorbance 2_01", "Abs", [.0866, .0902, .0877, *([None] * 9)]),
        ("Luminescence 1000_02", "RLU", [118.9, 152.6, 182.8, *([None] * 9)]),
    ]:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(["métadonnées variables"])
        sheet.append([None])
        sheet.append([marker, *range(1, 13)])
        sheet.append(["A", *([None] * 12)])
        sheet.append(["B", *values])
        for letter in "CDEFGH":
            sheet.append([letter, *([None] * 12)])
        sheet.append([])
        sheet.append(["Échantillon", *range(1, 13)])
        sheet.append(["A", *([None] * 12)])
        sheet.append(["B", *names])
        for letter in "CDEFGH":
            sheet.append([letter, *([None] * 12)])
    if mismatched_name:
        workbook["Luminescence 1000_02"].cell(15, 2).value = "Autre"
    if missing_lum:
        workbook["Luminescence 1000_02"].cell(5, 2).value = None
    plan = workbook.create_sheet("Plan de plaque")
    plan.append([None, *range(1, 13)])
    plan.append(["B", *names])
    plan.append([None, "Groupe 1", "Groupe 1", "Groupe 1"])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def test_single_time_parser_matches_abs_and_rlu_matrices_by_well():
    result = parse_single_time_workbook(single_time_workbook())
    assert result["puits"].tolist() == ["B01", "B02", "B03"]
    assert result["DO_brute"].tolist() == pytest.approx([.0866, .0902, .0877])
    assert result["Lum_brute"].tolist() == pytest.approx([118.9, 152.6, 182.8])
    assert result["sample_header"].tolist() == [
        "Blanc1 (B01)", "Blanc1 (B02)", "Blanc1 (B03)",
    ]
    assert result["replicat"].tolist() == [1, 2, 3]


def test_single_time_parser_rejects_name_disagreement_and_missing_measurement():
    with pytest.raises(ValueError, match="diffèrent pour le puits B01"):
        parse_single_time_workbook(single_time_workbook(mismatched_name=True))
    with pytest.raises(ValueError, match="n'a pas de valeur RLU"):
        parse_single_time_workbook(single_time_workbook(missing_lum=True))


def _single_time_table(time_hours: float) -> pd.DataFrame:
    table = parse_kinetic_workbook(synthetic_workbook())
    return table.loc[table["temps_h"].eq(time_hours)].reset_index(drop=True)


def _moving_well_table(sample_rows: str, sample_columns: range, time_index: int) -> pd.DataFrame:
    """Model the real 12-observation layout: three blanks plus three triplicates."""
    records = []
    conditions = [("Blanc", "Blanc", "SCFM2 (Po)")]
    wells_by_condition = [[f"B{column:02d}" for column in range(1, 4)]]
    for row, strain in zip(sample_rows, ("PspeD2-3B", "PAO1", "Kpne")):
        conditions.append(("Souche", strain, "SCFM2 (Po)"))
        wells_by_condition.append([f"{row}{column:02d}" for column in sample_columns])
    for (sample_type, strain, group), wells in zip(conditions, wells_by_condition):
        for replicate, well in enumerate(wells, start=1):
            physical_value = time_index * 1000 + ord(well[0]) * 10 + int(well[1:])
            records.append({
                "temps_h": 0.0, "souche": strain, "Groupe": group,
                "replicat": replicate, "DO_brute": physical_value / 1000,
                "Lum_brute": float(physical_value), "type": sample_type,
                "puits": well, "lecture": 1,
                "sample_header": f"{group} {strain} ({well})",
                "temps_sec_do": 0.0, "temps_sec_lum": 0.0,
                "ecart_temps_s": 0.0,
            })
    return pd.DataFrame(records)


def test_one_file_per_time_concatenates_and_maps_real_hours():
    result = combine_time_point_tables([
        ("experiment_t2.xlsx", _single_time_table(0)),
        ("experiment_t0.xlsx", _single_time_table(0)),
        ("experiment_t1.xlsx", _single_time_table(0)),
    ], {0: 0, 1: .5, 2: 3})
    assert result["time_index"].drop_duplicates().tolist() == [0, 1, 2]
    assert result.groupby("time_index")["temps_h"].first().to_dict() == {0: 0, 1: .5, 2: 3}
    assert result.groupby("time_index")["lecture"].first().to_dict() == {0: 1, 1: 2, 2: 3}
    assert result.groupby("time_index")["source_workbook"].first().to_dict() == {
        0: "experiment_t0.xlsx", 1: "experiment_t1.xlsx", 2: "experiment_t2.xlsx",
    }
    assert result.groupby("time_index")["temps_sec_do"].first().to_dict() == {
        0: 0, 1: 1800, 2: 10800,
    }
    assert result.groupby("time_index")["temps_sec_lum"].first().to_dict() == {
        0: 0, 1: 1800, 2: 10800,
    }
    assert result["experience"].unique().tolist() == ["experiment"]


def test_one_file_per_time_keeps_replicates_separate_and_accepts_gaps():
    result = combine_time_point_tables([
        ("rep1_t0.xlsx", _single_time_table(0)),
        ("rep1_t2.xlsx", _single_time_table(0)),
        ("rep2_t0.xlsx", _single_time_table(0)),
    ], {0: 0, 2: 6})
    assert set(result["experience"]) == {"rep1", "rep2"}
    assert result.groupby("experience")["Groupe"].first().nunique() == 2


def test_one_file_per_time_tracks_mobile_physical_wells_with_stable_curve_ids():
    tables = [
        ("real_t0.xlsx", _moving_well_table("FGH", range(10, 13), 0)),
        ("real_t1.xlsx", _moving_well_table("FGH", range(7, 10), 1)),
        ("real_t2.xlsx", _moving_well_table("FGH", range(4, 7), 2)),
    ]
    result = combine_time_point_tables(tables, {0: 0, 1: 1.5, 2: 4})

    assert result.groupby("time_index").size().to_dict() == {0: 12, 1: 12, 2: 12}
    assert result.groupby(["time_index", "souche"]).size().eq(3).all()
    assert result.groupby("sample_header")["time_index"].nunique().eq(3).all()
    pspe_rep1 = result.query("souche == 'PspeD2-3B' and replicat == 1")
    assert pspe_rep1["puits"].tolist() == ["F10", "F07", "F04"]
    assert pspe_rep1["source_sample_header"].tolist() == [
        "SCFM2 (Po) PspeD2-3B (F10)",
        "SCFM2 (Po) PspeD2-3B (F07)",
        "SCFM2 (Po) PspeD2-3B (F04)",
    ]
    assert pspe_rep1["Lum_brute"].tolist() == [710, 1707, 2704]
    assert pspe_rep1["DO_brute"].tolist() == pytest.approx([.710, 1.707, 2.704])
    assert result.groupby("time_index")["temps_h"].first().to_dict() == {0: 0, 1: 1.5, 2: 4}


def test_one_file_per_time_end_to_end_produces_visible_auc_figures():
    from luxplate.figure_lifecycle import validate_figure_render
    from luxplate.kinetics import run_kinetics
    from luxplate.plotting import build_publication_figures

    tables = [
        (f"real_t{index}.xlsx", _moving_well_table(rows, columns, index))
        for index, (rows, columns) in enumerate((
            ("FGH", range(10, 13)), ("CDE", range(7, 10)),
            ("HAB", range(4, 7)), ("DEF", range(1, 4)),
        ))
    ]
    combined = combine_time_point_tables(tables, {0: 0, 1: 1, 2: 2, 3: 3})
    # Stand in for the normal blank-correction stage while retaining the real
    # endpoint-import identities and moving physical-well audit columns.
    combined["DO_corr"] = .1 + .08 * combined["temps_h"] + .01 * combined["replicat"]
    combined["Lum_corr"] = 100 + 25 * combined["temps_h"] + combined["replicat"]
    combined["Lum_norm"] = combined["Lum_corr"] / combined["DO_corr"]
    combined["type"] = combined["type"].str.lower()

    result = run_kinetics(combined)

    assert len(result.series_metrics) == 9
    assert result.series_metrics.groupby("souche").size().eq(3).all()
    assert result.series_metrics["n_auc_points"].eq(4).all()
    assert result.series_metrics["lum_norm_auc"].notna().all()
    assert combined.query("type == 'souche'").groupby(
        ["souche", "replicat"]
    )["puits"].nunique().gt(1).all()
    figures = build_publication_figures(combined, families=("auc",))
    assert [name for name, _ in figures] == ["auc_luminescence_normalisee"]
    assert validate_figure_render(figures[0][1])["populated_axes"] > 0


def test_one_file_per_time_rejects_missing_biological_replicate():
    complete = _moving_well_table("FGH", range(10, 13), 0)
    incomplete = _moving_well_table("FGH", range(7, 10), 1)
    incomplete = incomplete.loc[~(
        incomplete["souche"].eq("PspeD2-3B") & incomplete["replicat"].eq(3)
    )]
    with pytest.raises(ValueError, match="structure biologique.*nombre de réplicats"):
        combine_time_point_tables([
            ("real_t0.xlsx", complete), ("real_t1.xlsx", incomplete),
        ], {0: 0, 1: 1})


def test_one_file_per_time_missing_index_is_reported_without_artificial_rows():
    names = [f"real_t{index}.xlsx" for index in (0, 1, 2, 3, 5)]
    _, missing = organize_time_files(names)
    tables = [(name, _moving_well_table("FGH", range(10, 13), index))
              for name, index in zip(names, (0, 1, 2, 3, 5))]
    result = combine_time_point_tables(tables, {index: float(index) for index in (0, 1, 2, 3, 5)})
    assert missing == {"real": [4]}
    assert result["time_index"].drop_duplicates().tolist() == [0, 1, 2, 3, 5]
    assert len(result) == 5 * 12


def test_one_file_per_time_requires_mapping_and_identical_biological_structure():
    first = _single_time_table(0)
    with pytest.raises(ValueError, match="n'est pas défini"):
        combine_time_point_tables([("run_t0.xlsx", first)], {})
    changed = first.loc[first["puits"].ne("A01")].copy()
    with pytest.raises(ValueError, match="structure biologique"):
        combine_time_point_tables([
            ("run_t0.xlsx", first), ("run_t1.xlsx", changed)
        ], {0: 0, 1: 1})


def test_per_time_and_legacy_modes_have_equivalent_internal_measurements():
    complete = parse_kinetic_workbook(synthetic_workbook())
    split = combine_time_point_tables([
        ("run_t0.xlsx", complete.query("temps_h == 0")),
        ("run_t1.xlsx", complete.query("temps_h == 1")),
    ], {0: 0, 1: 1})
    legacy = combine_kinetic_tables([("run.xlsx", complete)])
    scientific = [column for column in complete.columns if column not in {
        "temps_sec_do", "temps_sec_lum", "sample_header"
    }]
    order = ["type", "souche", "replicat", "temps_h"]
    pd.testing.assert_frame_equal(
        legacy.sort_values(order)[scientific].reset_index(drop=True),
        split.sort_values(order)[scientific].reset_index(drop=True),
    )


def workbook_with_strain_headers(headers: list[str]) -> BytesIO:
    """Build one plate whose headers are distinct technical wells."""
    if len(headers) > 96:
        raise ValueError("Une plaque synthétique ne peut pas dépasser 96 puits.")
    source = synthetic_workbook()
    workbook = load_workbook(source)
    wells = [f"{chr(ord('A') + index // 12)}{index % 12 + 1:02d}" for index in range(len(headers))]
    sample_headers = [f"{header} ({well})" for header, well in zip(headers, wells)]
    for sheet_name in ("Absorbance 600", "Luminescence 1"):
        sheet = workbook[sheet_name]
        sheet.delete_cols(3, sheet.max_column - 2)
        for column, header in enumerate(sample_headers, 3):
            sheet.cell(2, column).value = header
            sheet.cell(3, column).value = column
            sheet.cell(4, column).value = column * 2
    plan = workbook["Plan de plaque"]
    plan.delete_rows(5, max(plan.max_row - 4, 0))
    for row_index in range((len(headers) + 11) // 12):
        sample_row = 5 + row_index * 2
        plan.cell(sample_row, 1).value = chr(ord("A") + row_index)
        for column in range(2, min(14, len(headers) - row_index * 12 + 2)):
            plan.cell(sample_row, column).value = headers[row_index * 12 + column - 2]
            plan.cell(sample_row + 1, column).value = "Groupe 1"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def test_parse_kinetic_workbook_preserves_raw_technical_wells():
    result = parse_kinetic_workbook(synthetic_workbook())
    assert len(result) == 6
    assert result["sample_header"].nunique() == 3
    assert result.loc[result["souche"].eq("PAO1"), "replicat"].unique().tolist() == [1, 2]
    assert result.loc[result["souche"].eq("PAO1"), "Groupe"].eq("Groupe 1").all()
    np.testing.assert_allclose(sorted(result["temps_h"].unique()), [0, 1])
    assert result.loc[result["type"].eq("blanc"), "souche"].eq("Blanc").all()


def test_multiple_luminescence_sheets_require_explicit_selection():
    source = synthetic_workbook(second_luminescence=True)
    absorbance, luminescence = inspect_workbook(source)
    assert absorbance == "Absorbance 600"
    assert luminescence == ["Luminescence 1", "Luminescence 2"]
    source.seek(0)
    with pytest.raises(ValueError, match="choisissez-en une explicitement"):
        parse_kinetic_workbook(source)
    source.seek(0)
    selected = parse_kinetic_workbook(source, "Luminescence 2")
    assert selected["Lum_brute"].max() == 40


def test_medium_prefix_is_separated_and_numbered_blank_is_associated():
    source = synthetic_workbook()
    workbook = load_workbook(source)
    renamed = {
        "PAO1 (A01)": "SCFM2 (Po) 14.1Ac attB::P0-lux (A01)",
        "PAO1 (A02)": "SCFM2 (Po) 14.1Ac attB::P0-lux (A02)",
        "Blanc (B01)": "Blanc1 (B01)",
    }
    for sheet_name in ("Absorbance 600", "Luminescence 1"):
        sheet = workbook[sheet_name]
        for cell in sheet[2]:
            if cell.value in renamed:
                cell.value = renamed[cell.value]
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    result = parse_kinetic_workbook(output)

    strains = result.loc[result["type"].eq("souche")]
    blanks = result.loc[result["type"].eq("blanc")]
    assert strains["souche"].unique().tolist() == ["14.1Ac attB::P0-lux"]
    assert strains["Groupe"].unique().tolist() == ["SCFM2 (Po)"]
    assert blanks["Groupe"].unique().tolist() == ["SCFM2 (Po)"]


@pytest.mark.parametrize(
    ("header", "expected"),
    [
        ("14.1Ac attB::MiniCTXlux(PspeD2-1A-lux) (A01)", "14.1Ac attB::PspeD2-1A-lux"),
        ("14.1Ac attB::PspeD2-1A-Lux (A01)", "14.1Ac attB::PspeD2-1A-lux"),
        ("14.1Ac attB::PspeD2-1A-lux (A01)", "14.1Ac attB::PspeD2-1A-lux"),
        ("14.1Ac attB::MiniCTXlux(P0-lux) (A01)", "14.1Ac attB::P0-lux"),
    ],
)
def test_equivalent_lux_reporter_spellings_share_a_canonical_strain(header, expected):
    source = synthetic_workbook()
    workbook = load_workbook(source)
    for sheet_name in ("Absorbance 600", "Luminescence 1"):
        sheet = workbook[sheet_name]
        sheet.cell(2, 3).value = header
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    result = parse_kinetic_workbook(output)

    assert result.loc[result["puits"].eq("A01"), "souche"].unique().tolist() == [expected]


def test_varioskan_0001_suffix_is_removed_without_losing_biological_qualifiers():
    assert normalize_strain_name("14.1Ac attB::P0-lux0001") == normalize_strain_name(
        "14.1Ac attB::P0-Lux"
    )
    assert normalize_strain_name("14.1Ac attB::PspeD2-1A-Lux") == normalize_strain_name(
        "14.1Ac attB::PspeD2-1A-lux"
    )
    assert normalize_strain_name("14.1Ac attB::PspeD2-3B-Lux") == normalize_strain_name(
        "14.1Ac attB::PspeD2-3B-lux"
    )
    assert normalize_strain_name("14.1Ac attB::P0-lux (M1)0001") == (
        "14.1Ac attB::P0-lux (M1)"
    )
    assert normalize_strain_name("14.1Ac attB::P0-lux (M1)0001") != normalize_strain_name(
        "14.1Ac attB::P0-lux (M2)0001"
    )


@pytest.mark.parametrize("attachment_site", ["attb", "attB", "ATTB"])
def test_attachment_site_case_variants_use_canonical_attB_spelling(attachment_site):
    assert normalize_strain_name(f"14.1Ac {attachment_site}::P0-lux") == (
        "14.1Ac attB::P0-lux"
    )


def test_attachment_site_case_variants_collapse_to_one_strain_during_import():
    headers = ["14.1Ac attB::P0-lux", "14.1Ac attb::P0-lux"]

    strains = parse_kinetic_workbook(workbook_with_strain_headers(headers)).query("type == 'souche'")

    assert strains["souche"].unique().tolist() == ["14.1Ac attB::P0-lux"]
    assert strains["replicat"].unique().tolist() == [1, 2]


def test_lone_varioskan_0001_header_is_canonicalized_during_import():
    headers = ["LB 14.1Ac attB::P0-lux0001"] * 3

    strains = parse_kinetic_workbook(workbook_with_strain_headers(headers)).query("type == 'souche'")

    assert strains["souche"].unique().tolist() == ["14.1Ac attB::P0-lux"]
    assert strains["Groupe"].unique().tolist() == ["LB"]
    assert strains["replicat"].unique().tolist() == [1, 2, 3]
    assert strains["puits"].nunique() == 3


def test_morphotypes_remain_distinct_during_import():
    headers = [
        "SCFM2-KPI 14.1Ac attB::P0-lux (M1)0001",
        "SCFM2-KPI 14.1Ac attB::P0-lux (M2)0001",
    ]

    strains = parse_kinetic_workbook(workbook_with_strain_headers(headers)).query("type == 'souche'")

    assert strains["souche"].unique().tolist() == [
        "14.1Ac attB::P0-lux (M1)",
        "14.1Ac attB::P0-lux (M2)",
    ]
    assert strains["Groupe"].unique().tolist() == ["SCFM2-KPI"]


@pytest.mark.parametrize(
    "headers",
    [
        ["P0-lux", "P0-lux0002", "P0-lux0003"],
        ["PspeD2-1A-lux", "PspeD2-1A-lux0002"],
        ["PspeD2-3B-lux", "PspeD2-3B-lux0025"],
        ["P0-lux", "P0-lux2", "P0-lux_2", "P0-lux-2", "P0-lux_rep2", "P0-lux-rep2"],
    ],
)
def test_source_generated_technical_suffixes_share_one_strain(headers):
    result = parse_kinetic_workbook(workbook_with_strain_headers(headers))

    strains = result.loc[result["type"].eq("souche")]
    assert strains["souche"].unique().tolist() == [headers[0]]
    assert strains["replicat"].unique().tolist() == list(range(1, len(headers) + 1))
    assert strains["sample_header"].nunique() == len(headers)
    assert strains["puits"].nunique() == len(headers)


@pytest.mark.parametrize("promoter", ["PspeD2-1A", "PspeD2-3B"])
def test_numbered_lux_series_is_collapsed_without_unsuffixed_header(promoter):
    headers = [f"{promoter}-lux{number:04d}" for number in (1, 2, 3, 25)]

    strains = parse_kinetic_workbook(workbook_with_strain_headers(headers)).query("type == 'souche'")

    assert strains["souche"].unique().tolist() == [f"{promoter}-lux"]
    assert strains["replicat"].unique().tolist() == [1, 2, 3, 4]
    assert strains["sample_header"].nunique() == 4
    assert strains["puits"].nunique() == 4


@pytest.mark.parametrize("promoter", ["PspeD2-1A", "PspeD2-3B"])
def test_complete_numbered_lux_series_has_25_replicates(promoter):
    headers = [f"{promoter}-lux{number:04d}" for number in range(1, 26)]

    strains = parse_kinetic_workbook(workbook_with_strain_headers(headers)).query("type == 'souche'")

    assert strains["souche"].unique().tolist() == [f"{promoter}-lux"]
    assert strains["replicat"].unique().tolist() == list(range(1, 26))


def test_ma_complex_reference_structure_has_exactly_three_strains():
    strain_prefix = "14.1Ac attb::"
    headers = (
        [f"{strain_prefix}P0-lux"] * 22
        + [f"{strain_prefix}PspeD2-1A-lux{number:04d}" for number in range(1, 26)]
        + [f"{strain_prefix}PspeD2-3B-lux{number:04d}" for number in range(1, 26)]
    )

    strains = parse_kinetic_workbook(workbook_with_strain_headers(headers)).query("type == 'souche'")

    canonical_prefix = "14.1Ac attB::"
    assert strains["souche"].unique().tolist() == [
        f"{canonical_prefix}P0-lux",
        f"{canonical_prefix}PspeD2-1A-lux",
        f"{canonical_prefix}PspeD2-3B-lux",
    ]
    assert strains.groupby("souche")["sample_header"].nunique().to_dict() == {
        f"{canonical_prefix}P0-lux": 22,
        f"{canonical_prefix}PspeD2-1A-lux": 25,
        f"{canonical_prefix}PspeD2-3B-lux": 25,
    }


@pytest.mark.parametrize("natural_name", ["PspeD2", "14.1Ac", "14.3B", "PA14", "3-B", "1-A"])
def test_natural_terminal_digits_are_not_stripped_without_cohort_evidence(natural_name):
    result = parse_kinetic_workbook(workbook_with_strain_headers([natural_name]))

    assert result.loc[result["type"].eq("souche"), "souche"].unique().tolist() == [natural_name]


def test_combining_workbooks_namespaces_internal_blank_links():
    first = parse_kinetic_workbook(synthetic_workbook())
    second = parse_kinetic_workbook(synthetic_workbook())

    combined = combine_kinetic_tables([("jour1.xlsx", first), ("jour2.xlsx", second)])

    assert set(combined["experience"]) == {"jour1", "jour2"}
    assert combined["Groupe"].nunique() == 2
    assert combined["sample_header"].nunique() == first["sample_header"].nunique() * 2
