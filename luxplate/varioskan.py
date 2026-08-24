"""Import Varioskan kinetic workbooks into the historical long-table schema.

This module is the application-facing integration of
``examples/synthetic/01_mise_en_forme_donnees.py``.  It deliberately keeps the
raw values and the French column names used by the subsequent legacy stages.
"""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import BinaryIO, Iterable

import pandas as pd
from openpyxl import load_workbook

HEADER_READING = "Lecture en cours"
HEADER_TIME = "temps moy. [s]"
PLATE_ROWS = set("ABCDEFGH")
STRAIN_START = re.compile(
    # Common P. aeruginosa strain names used in LuxPlate exports.  A medium
    # may contain spaces/parentheses, so look for the strain token rather than
    # trying to split the label on whitespace.
    r"(?<!\S)(?=(?:\d+(?:\.\d+)+[A-Za-z][A-Za-z0-9.-]*|PA(?:O)?\d+[A-Za-z0-9.-]*)(?:\s|$))",
    flags=re.IGNORECASE,
)
BLANK_NUMBER = re.compile(r"^(?:blanc|blank)\s*[-_ ]?(\d+)\b", flags=re.IGNORECASE)
MINICTX_REPORTER = re.compile(r"^MiniCTXlux\s*\((.+)\)$", flags=re.IGNORECASE)
# Varioskan can append ``0001`` to an exported sample label, including after a
# biologically meaningful qualifier such as ``(M1)``.  Restrict the rule to a
# lux reporter identity: stripping arbitrary terminal digits would corrupt
# real strain/construct names such as PA14, PspeD2, or PspeD2-1A.
VARIOSKAN_LUX_0001_SUFFIX = re.compile(
    r"(?P<identity>-lux(?:\s+\([^()]+\))?)0001$",
    flags=re.IGNORECASE,
)
TECHNICAL_SUFFIXES = (
    # Explicit replicate markers are unambiguous, but are still only removed
    # when the corresponding unsuffixed sample occurs in the same import.
    re.compile(r"(?:[_-]?rep(?:licat|licate)?[_-]?\d+)$", re.IGNORECASE),
    re.compile(r"(?:[_-]\d+)$"),
    re.compile(r"(?:\d+)$"),
)
NUMBERED_LUX_SERIES = re.compile(r"^(.*-lux)(\d{4})$", flags=re.IGNORECASE)
BIOLOGICAL_REPLICATE_MARKER = re.compile(
    r"(?:^|[_\-\s])rep(?:licat|licate)?[_\-\s]*0*(\d+)(?:$|[_\-\s])",
    flags=re.IGNORECASE,
)


def clean_text(value: object) -> str:
    """Normalize spreadsheet text without changing meaningful labels."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def normalize_strain_name(value: object) -> str:
    """Return a canonical name for known equivalent lux reporter spellings.

    Varioskan headers are entered manually and the same chromosomal reporter is
    commonly written either as ``MiniCTXlux(PspeD2-1A-lux)`` or directly as
    ``PspeD2-1A-Lux``.  The vector wrapper is not part of the strain identity in
    this workflow, and ``lux`` capitalization is likewise insignificant.
    """
    name = clean_text(value)
    prefix, separator, construct = name.rpartition("::")
    if not separator:
        prefix, construct = "", name
    else:
        # attB is the conventional spelling of the chromosomal attachment
        # site.  Treat manually entered case variants (notably ``attb``) as
        # the same strain so they do not appear as duplicate UI choices.
        prefix = re.sub(r"\battb\b", "attB", prefix, flags=re.IGNORECASE)

    wrapped = MINICTX_REPORTER.fullmatch(construct)
    if wrapped:
        construct = clean_text(wrapped.group(1))
    construct = VARIOSKAN_LUX_0001_SUFFIX.sub(r"\g<identity>", construct)
    # Canonicalize the reporter token even when a meaningful qualifier follows
    # it (for example ``P0-Lux (M1)``), rather than only at end of string.
    construct = re.sub(r"-lux(?=$|\s+\()", "-lux", construct, flags=re.IGNORECASE)
    return f"{prefix}::{construct}" if separator else construct


def _canonicalize_technical_suffixes(metadata: pd.DataFrame) -> pd.DataFrame:
    """Collapse source-generated replicate suffixes using cohort evidence.

    A trailing number is intrinsically ambiguous in a biological name.  A
    candidate without the suffix is normally accepted only when that base name
    also occurs in the same plate group.  Varioskan additionally emits lux
    technical series as ``-lux0001``, ``-lux0002``, ... without emitting an
    unsuffixed header.  Two or more distinct four-digit members provide enough
    cohort evidence to collapse that series safely.  The well/header remain
    untouched and ``replicat`` is assigned later, so no technical series is
    lost or promoted to a biological replicate.
    """
    output = metadata.copy()
    strains = output["type"].eq("souche")
    for _, indices in output.loc[strains].groupby("Groupe", sort=False).groups.items():
        observed = set(output.loc[indices, "souche"])
        replacements: dict[str, str] = {}

        numbered_lux: dict[str, set[str]] = {}
        for name in observed:
            match = NUMBERED_LUX_SERIES.fullmatch(name)
            if match:
                numbered_lux.setdefault(match.group(1), set()).add(name)
        for base, members in numbered_lux.items():
            if len(members) >= 2:
                replacements.update(dict.fromkeys(members, base))

        for name in observed:
            candidates = []
            for suffix in TECHNICAL_SUFFIXES:
                match = suffix.search(name)
                if match:
                    candidates.append(clean_text(name[:match.start()]))
            # Prefer the longest observed prefix.  This matters when a real
            # construct itself ends in digits (for example PA14).
            valid = [candidate for candidate in candidates if candidate in observed]
            if valid and name not in replacements:
                replacements[name] = max(valid, key=len)
        output.loc[indices, "souche"] = output.loc[indices, "souche"].replace(replacements)
    return output


def find_measurement_sheets(sheet_names: Iterable[str]) -> tuple[str, list[str]]:
    """Return the first absorbance sheet and every luminescence candidate."""
    names = list(sheet_names)
    absorbance = [n for n in names if clean_text(n).lower().startswith("absorbance")]
    luminescence = [n for n in names if clean_text(n).lower().startswith("luminescence")]
    if not absorbance:
        raise ValueError("Aucune feuille d'absorbance détectée dans le fichier.")
    if not luminescence:
        raise ValueError("Aucune feuille de luminescence détectée dans le fichier.")
    return absorbance[0], luminescence


def inspect_workbook(source: str | Path | BinaryIO) -> tuple[str, list[str]]:
    """Inspect a workbook so the UI can require an explicit Lum sheet choice."""
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        return find_measurement_sheets(workbook.sheetnames)
    finally:
        workbook.close()


def _header_row(sheet) -> int:
    for row in range(1, min(sheet.max_row, 30) + 1):
        values = [clean_text(sheet.cell(row, col).value) for col in range(1, min(sheet.max_column, 10) + 1)]
        if HEADER_READING in values and HEADER_TIME in values:
            return row
    raise ValueError(f"Impossible de trouver la ligne d'entête dans la feuille {sheet.title!r}.")


def _wide_table(sheet) -> tuple[pd.DataFrame, list[str]]:
    header_row = _header_row(sheet)
    headers = [clean_text(sheet.cell(header_row, col).value) for col in range(1, sheet.max_column + 1)]
    useful = [index for index, name in enumerate(headers, start=1) if name]
    rows = []
    for row in range(header_row + 1, sheet.max_row + 1):
        values = [sheet.cell(row, col).value for col in useful]
        if not values or all(value is None for value in values):
            continue
        rows.append(values)
    if not rows:
        raise ValueError(f"Aucune donnée détectée dans {sheet.title!r}.")

    frame = pd.DataFrame(rows, columns=[headers[index - 1] for index in useful]).rename(
        columns={HEADER_READING: "lecture", HEADER_TIME: "temps_sec"}
    )
    frame["lecture"] = pd.to_numeric(frame["lecture"], errors="coerce")
    frame["temps_sec"] = pd.to_numeric(frame["temps_sec"], errors="coerce")
    frame = frame.dropna(subset=["lecture", "temps_sec"]).copy()
    frame["lecture"] = frame["lecture"].astype(int)
    measurements = [column for column in frame.columns if column not in {"lecture", "temps_sec"}]
    for column in measurements:
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    return frame, measurements


def _sample_metadata(header: str) -> dict[str, str]:
    normalized = clean_text(header)
    match = re.match(r"^(.*?)(?:\s*\(([A-H]\d{2})\))?$", normalized)
    name = clean_text(match.group(1)) if match else normalized
    well = (match.group(2) or "") if match else ""
    kind = "blanc" if re.search(r"\b(?:blanc|blank)\s*\d*\b", name, re.IGNORECASE) else "souche"
    medium = ""
    strain = name
    if kind == "souche":
        strain_match = STRAIN_START.search(name)
        if strain_match and strain_match.start() > 0:
            medium = clean_text(name[:strain_match.start()])
            strain = clean_text(name[strain_match.start():])
        strain = normalize_strain_name(strain)
    return {
        "sample_header": header,
        "souche": strain,
        "puits": well,
        "type": kind,
        "milieu_entete": medium,
        "nom_echantillon": name,
    }


def _resolve_groups(metadata: pd.DataFrame, plate_groups: dict[str, str]) -> pd.Series:
    """Prefer media encoded before strain names and associate numbered blanks.

    In real exports ``Blanc1`` commonly denotes the blank for the first medium,
    while the plate-plan group is merely a technical block (``Groupe 1``).
    """
    groups = metadata["puits"].map(plate_groups).fillna("")
    strain_media = list(dict.fromkeys(
        metadata.loc[metadata["type"].eq("souche"), "milieu_entete"].loc[lambda values: values.ne("")]
    ))
    encoded = metadata["milieu_entete"].ne("")
    groups.loc[encoded] = metadata.loc[encoded, "milieu_entete"]

    for index, row in metadata.loc[metadata["type"].eq("blanc")].iterrows():
        number = BLANK_NUMBER.match(row["nom_echantillon"])
        if number and 1 <= int(number.group(1)) <= len(strain_media):
            groups.loc[index] = strain_media[int(number.group(1)) - 1]
        elif len(strain_media) == 1:
            groups.loc[index] = strain_media[0]
    return groups


def _plate_groups(sheet) -> dict[str, str]:
    groups: dict[str, str] = {}
    number_row = None
    for row in range(1, min(sheet.max_row, 20) + 1):
        numbers = set()
        for col in range(2, min(sheet.max_column, 13) + 1):
            try:
                numbers.add(int(sheet.cell(row, col).value))
            except (TypeError, ValueError):
                pass
        if set(range(1, 13)).issubset(numbers):
            number_row = row
            break
    if number_row is None:
        raise ValueError("Impossible de trouver l'entête 1..12 dans le Plan de plaque.")

    for row in range(1, sheet.max_row + 1):
        letter = clean_text(sheet.cell(row, 1).value).upper()
        if letter not in PLATE_ROWS or row + 1 > sheet.max_row:
            continue
        for col in range(2, min(sheet.max_column, 13) + 1):
            try:
                number = int(sheet.cell(number_row, col).value)
            except (TypeError, ValueError):
                continue
            group = clean_text(sheet.cell(row + 1, col).value)
            if group:
                groups[f"{letter}{number:02d}"] = group
    return groups


def _matrix_marker(value: object) -> str:
    """Normalize a Varioskan matrix title for language-independent matching."""
    text = unicodedata.normalize("NFKD", clean_text(value))
    return "".join(character for character in text if not unicodedata.combining(character)).casefold()


def _plate_matrix(sheet, markers: set[str], description: str) -> dict[str, object]:
    """Find and read an A-H by 1-12 matrix without relying on export row offsets."""
    normalized_markers = {_matrix_marker(marker) for marker in markers}
    header_row = None
    number_columns: dict[int, int] = {}
    for row in range(1, sheet.max_row + 1):
        if _matrix_marker(sheet.cell(row, 1).value) not in normalized_markers:
            continue
        candidate: dict[int, int] = {}
        for column in range(2, sheet.max_column + 1):
            try:
                number = int(clean_text(sheet.cell(row, column).value))
            except (TypeError, ValueError):
                continue
            if 1 <= number <= 12 and number not in candidate:
                candidate[number] = column
        if set(candidate) == set(range(1, 13)):
            header_row, number_columns = row, candidate
            break
    if header_row is None:
        raise ValueError(
            f"Impossible de trouver la matrice {description} (entête 1..12) "
            f"dans la feuille {sheet.title!r}."
        )

    values: dict[str, object] = {}
    found_rows: set[str] = set()
    for row in range(header_row + 1, sheet.max_row + 1):
        letter = clean_text(sheet.cell(row, 1).value).upper()
        if letter in PLATE_ROWS and letter not in found_rows:
            found_rows.add(letter)
            for number, column in number_columns.items():
                values[f"{letter}{number:02d}"] = sheet.cell(row, column).value
            if found_rows == PLATE_ROWS:
                break
        elif found_rows and _matrix_marker(sheet.cell(row, 1).value) in {
            "abs", "absorbance", "rlu", "echantillon", "sample", "samples"
        }:
            break
    if not found_rows:
        raise ValueError(f"Aucune ligne A..H détectée dans la matrice {description} de {sheet.title!r}.")
    return values


def parse_single_time_workbook(
    source: str | Path | BinaryIO,
    luminescence_sheet: str | None = None,
) -> pd.DataFrame:
    """Parse one endpoint-style Varioskan workbook as a single time point.

    Unlike :func:`parse_kinetic_workbook`, this parser consumes the plate
    matrices headed by ``Abs``, ``RLU`` and ``Échantillon``.  Both paths emit
    the same long-table schema before entering the scientific pipeline.
    """
    workbook = load_workbook(source, read_only=False, data_only=True)
    try:
        absorbance_sheet, luminescence_sheets = find_measurement_sheets(workbook.sheetnames)
        if luminescence_sheet is None:
            if len(luminescence_sheets) > 1:
                raise ValueError("Plusieurs feuilles de luminescence détectées : choisissez-en une explicitement.")
            luminescence_sheet = luminescence_sheets[0]
        if luminescence_sheet not in luminescence_sheets:
            raise ValueError(f"Feuille de luminescence inconnue : {luminescence_sheet!r}.")
        plan_names = [name for name in workbook.sheetnames if _matrix_marker(name) == "plan de plaque"]
        if not plan_names:
            raise ValueError("La feuille 'Plan de plaque' est absente.")

        absorbance = _plate_matrix(workbook[absorbance_sheet], {"Abs", "Absorbance"}, "d'absorbance")
        luminescence = _plate_matrix(workbook[luminescence_sheet], {"RLU"}, "de luminescence")
        sample_markers = {"Échantillon", "Echantillon", "Sample", "Samples"}
        absorbance_names = _plate_matrix(workbook[absorbance_sheet], sample_markers, "des échantillons")
        luminescence_names = _plate_matrix(workbook[luminescence_sheet], sample_markers, "des échantillons")

        named_wells = []
        for well in sorted(set(absorbance_names) | set(luminescence_names)):
            do_name = clean_text(absorbance_names.get(well))
            lum_name = clean_text(luminescence_names.get(well))
            if do_name != lum_name:
                raise ValueError(
                    f"Les noms d'échantillon diffèrent pour le puits {well} : "
                    f"absorbance={do_name!r}, luminescence={lum_name!r}."
                )
            if do_name:
                named_wells.append((well, do_name))
        if not named_wells:
            raise ValueError("Aucun puits nommé n'a été détecté dans les matrices Échantillon.")

        rows = []
        for well, name in named_wells:
            do_value = pd.to_numeric(absorbance.get(well), errors="coerce")
            lum_value = pd.to_numeric(luminescence.get(well), errors="coerce")
            if pd.isna(do_value) or pd.isna(lum_value):
                missing = "DO" if pd.isna(do_value) else "RLU"
                raise ValueError(f"Le puits nommé {well} ({name}) n'a pas de valeur {missing} exploitable.")
            header = f"{name} ({well})"
            rows.append({**_sample_metadata(header), "DO_brute": float(do_value),
                         "Lum_brute": float(lum_value)})

        metadata = pd.DataFrame(rows)
        metadata["Groupe"] = _resolve_groups(metadata, _plate_groups(workbook[plan_names[0]]))
        metadata = _canonicalize_technical_suffixes(metadata)
        metadata["replicat"] = metadata.groupby(["souche", "Groupe"], sort=False).cumcount() + 1
        metadata["temps_h"] = 0.0
        metadata["lecture"] = 1
        metadata["temps_sec_do"] = 0.0
        metadata["temps_sec_lum"] = 0.0
        metadata["ecart_temps_s"] = 0.0
        columns = [
            "temps_h", "souche", "Groupe", "replicat", "DO_brute", "Lum_brute",
            "type", "puits", "lecture", "sample_header", "temps_sec_do",
            "temps_sec_lum", "ecart_temps_s",
        ]
        return metadata[columns].sort_values(["type", "souche", "replicat"]).reset_index(drop=True)
    finally:
        workbook.close()


def parse_kinetic_workbook(
    source: str | Path | BinaryIO,
    luminescence_sheet: str | None = None,
) -> pd.DataFrame:
    """Parse one kinetic workbook into an unaveraged, raw long table.

    When multiple luminescence sheets exist, callers must select one rather
    than allowing the scientific choice to be made silently.
    """
    workbook = load_workbook(source, read_only=False, data_only=True)
    try:
        absorbance_sheet, luminescence_sheets = find_measurement_sheets(workbook.sheetnames)
        if luminescence_sheet is None:
            if len(luminescence_sheets) > 1:
                raise ValueError("Plusieurs feuilles de luminescence détectées : choisissez-en une explicitement.")
            luminescence_sheet = luminescence_sheets[0]
        if luminescence_sheet not in luminescence_sheets:
            raise ValueError(f"Feuille de luminescence inconnue : {luminescence_sheet!r}.")
        plan_names = [name for name in workbook.sheetnames if clean_text(name).lower() == "plan de plaque"]
        if not plan_names:
            raise ValueError("La feuille 'Plan de plaque' est absente.")

        absorbance, absorbance_columns = _wide_table(workbook[absorbance_sheet])
        luminescence, luminescence_columns = _wide_table(workbook[luminescence_sheet])
        common = [column for column in absorbance_columns if column in set(luminescence_columns)]
        if not common:
            raise ValueError("Aucun échantillon commun entre l'absorbance et la luminescence.")

        do_long = absorbance[["lecture", "temps_sec", *common]].melt(
            id_vars=["lecture", "temps_sec"], var_name="sample_header", value_name="DO_brute"
        ).rename(columns={"temps_sec": "temps_sec_do"})
        lum_long = luminescence[["lecture", "temps_sec", *common]].melt(
            id_vars=["lecture", "temps_sec"], var_name="sample_header", value_name="Lum_brute"
        ).rename(columns={"temps_sec": "temps_sec_lum"})
        result = do_long.merge(lum_long, on=["lecture", "sample_header"], how="inner", validate="one_to_one")
        result["ecart_temps_s"] = (result["temps_sec_do"] - result["temps_sec_lum"]).abs()
        result["temps_h"] = result[["temps_sec_do", "temps_sec_lum"]].mean(axis=1) / 3600.0

        metadata = pd.DataFrame([_sample_metadata(column) for column in common])
        metadata["Groupe"] = _resolve_groups(metadata, _plate_groups(workbook[plan_names[0]]))
        metadata = _canonicalize_technical_suffixes(metadata)
        metadata["replicat"] = metadata.groupby(["souche", "Groupe"], sort=False).cumcount() + 1
        result = result.merge(metadata, on="sample_header", how="left", validate="many_to_one")
        columns = [
            "temps_h", "souche", "Groupe", "replicat", "DO_brute", "Lum_brute",
            "type", "puits", "lecture", "sample_header", "temps_sec_do",
            "temps_sec_lum", "ecart_temps_s",
        ]
        return result[columns].sort_values(["type", "souche", "replicat", "temps_h"]).reset_index(drop=True)
    finally:
        workbook.close()


def combine_kinetic_tables(
    tables: Iterable[tuple[str, pd.DataFrame]],
) -> pd.DataFrame:
    """Combine several parsed workbooks without mixing their technical groups.

    Plate-plan groups are an internal blank-association key.  Namespacing that
    key (and sample headers) by workbook prevents identically named wells or
    groups from unrelated experiments from being merged during QC/correction.
    The original workbook name remains available in ``experience`` for exports.
    """
    combined: list[pd.DataFrame] = []
    for position, (name, table) in enumerate(tables, start=1):
        frame = table.copy(deep=True)
        experience = clean_text(Path(name).stem) or f"experience_{position}"
        namespace = f"exp{position}"
        frame["experience"] = experience
        frame["Groupe"] = namespace + "|" + frame["Groupe"].fillna("").astype(str)
        frame["sample_header"] = namespace + "|" + frame["sample_header"].fillna("").astype(str)
        combined.append(frame)
    if not combined:
        raise ValueError("Ajoutez au moins un classeur à analyser.")
    return pd.concat(combined, ignore_index=True, sort=False)


TIME_FILE_RE = re.compile(r"_t(?P<index>\d+)(?=$|[(_\-.])", re.IGNORECASE)


def parse_time_file_name(name: str) -> tuple[str, int]:
    """Return the experiment stem and numeric time point encoded in ``name``.

    Text after the marker is deliberately ignored, so ``run_t2_valid.xlsx``
    conflicts with ``run_t2.xlsx`` instead of silently becoming another run.
    """
    stem = Path(name).stem
    matches = list(TIME_FILE_RE.finditer(stem))
    if not matches:
        raise ValueError(f"Le fichier {name!r} ne contient pas de point '_tX'.")
    match = matches[-1]
    experiment = clean_text(stem[:match.start()])
    if not experiment:
        raise ValueError(f"Le fichier {name!r} n'a pas d'identifiant d'expérience avant '_tX'.")
    return experiment, int(match.group("index"))


def organize_time_files(names: Iterable[str]) -> tuple[dict[str, list[tuple[int, str]]], dict[str, list[int]]]:
    """Group names by experiment, reject duplicates, and report index gaps."""
    groups: dict[str, list[tuple[int, str]]] = {}
    seen: dict[tuple[str, int], list[str]] = {}
    for name in names:
        experiment, index = parse_time_file_name(name)
        groups.setdefault(experiment, []).append((index, name))
        seen.setdefault((experiment, index), []).append(name)
    conflicts = [(key, files) for key, files in seen.items() if len(files) > 1]
    if conflicts:
        experiment, index = conflicts[0][0]
        files = "\n- ".join(conflicts[0][1])
        raise ValueError(
            f"Deux fichiers correspondent au point t{index} de l'expérience {experiment} :\n"
            f"- {files}\nVeuillez conserver un seul fichier pour ce point temporel."
        )
    missing: dict[str, list[int]] = {}
    for experiment, files in groups.items():
        files.sort(key=lambda item: item[0])
        indices = [item[0] for item in files]
        missing[experiment] = sorted(set(range(min(indices), max(indices) + 1)) - set(indices))
    return groups, missing


def suggest_regular_time_mapping(
    indices: Iterable[int], *, first_time: float = 0.0, interval: float = 1.0,
) -> dict[int, float]:
    """Propose real times for file indices sampled at a regular interval.

    The numeric ``tX`` value determines the position in the series.  This also
    preserves the intended elapsed time when an intermediate file is absent.
    """
    points = sorted(set(indices))
    if not points:
        return {}
    if interval <= 0:
        raise ValueError("L'intervalle entre deux points doit être strictement positif.")
    origin = points[0]
    return {
        index: float(first_time + (index - origin) * interval)
        for index in points
    }


def _plate_signature(table: pd.DataFrame) -> pd.DataFrame:
    required = {"sample_header", "puits", "type", "souche", "Groupe", "replicat"}
    absent = sorted(required - set(table.columns))
    if absent:
        raise ValueError(f"Colonnes nécessaires absentes : {', '.join(absent)}.")
    return (table[list(sorted(required))].drop_duplicates()
            .sort_values(list(sorted(required))).reset_index(drop=True))


def combine_time_point_tables(
    tables: Iterable[tuple[str, pd.DataFrame]], time_mapping: dict[int, float],
) -> pd.DataFrame:
    """Normalize one-workbook-per-time input to Lum's standard long table."""
    items = list(tables)
    groups, _ = organize_time_files(name for name, _ in items)
    by_name = dict(items)
    combined: list[pd.DataFrame] = []
    for position, (experiment, files) in enumerate(groups.items(), start=1):
        reference = None
        for index, name in files:
            if index not in time_mapping or pd.isna(time_mapping[index]):
                raise ValueError(f"Le temps réel du point t{index} n'est pas défini.")
            frame = by_name[name]
            if frame["temps_h"].nunique(dropna=False) != 1:
                raise ValueError(f"Le fichier {name!r} doit contenir un seul temps de mesure.")
            signature = _plate_signature(frame)
            if reference is None:
                reference = signature
            elif not signature.equals(reference):
                raise ValueError(
                    f"La structure de plaque ou le nombre de puits diffère dans {name!r} "
                    f"pour l'expérience {experiment}."
                )
            normalized = frame.copy(deep=True)
            normalized["time_index"] = index
            normalized["temps_h"] = float(time_mapping[index])
            normalized["temps_sec_do"] = normalized["temps_h"] * 3600.0
            normalized["temps_sec_lum"] = normalized["temps_h"] * 3600.0
            normalized["ecart_temps_s"] = 0
            normalized["lecture"] = index + 1
            normalized["experience"] = experiment
            namespace = f"exp{position}|"
            normalized["Groupe"] = namespace + normalized["Groupe"].fillna("").astype(str)
            normalized["sample_header"] = namespace + normalized["sample_header"].fillna("").astype(str)
            combined.append(normalized)
    if not combined:
        raise ValueError("Ajoutez au moins un classeur à analyser.")
    return pd.concat(combined, ignore_index=True, sort=False).sort_values(
        ["experience", "time_index", "type", "souche", "replicat"]
    ).reset_index(drop=True)


def suggest_biological_pair_id(experience: object) -> str:
    """Suggest an editable cross-file biological identity from a run name.

    This is deliberately only a UI proposal.  In particular, file order and
    equal cohort sizes are never used to manufacture pairs.  An unnumbered run
    is proposed as the conventional first replicate, while explicit ``repN``
    markers retain their number.
    """
    match = BIOLOGICAL_REPLICATE_MARKER.search(clean_text(experience))
    return f"bio{int(match.group(1))}" if match else "bio1"


def assign_biological_pair_ids(
    data: pd.DataFrame, mapping: pd.DataFrame,
) -> pd.DataFrame:
    """Attach user-validated biological pair identities to imported rows."""
    required = {"experience", "biological_pair_id"}
    if not required.issubset(mapping.columns) or "experience" not in data:
        raise ValueError("La table d'appariement biologique est incomplète.")
    pairs = mapping[list(required)].copy()
    pairs["experience"] = pairs["experience"].map(clean_text)
    pairs["biological_pair_id"] = pairs["biological_pair_id"].map(clean_text)
    if pairs["experience"].duplicated().any():
        raise ValueError("Chaque fichier doit apparaître une seule fois dans la table d'appariement.")
    expected = set(data["experience"].map(clean_text).unique())
    if set(pairs["experience"]) != expected or pairs["biological_pair_id"].eq("").any():
        raise ValueError("Renseignez un réplicat biologique non vide pour chaque fichier.")
    result = data.drop(columns="biological_pair_id", errors="ignore").copy()
    return result.merge(pairs, on="experience", how="left", validate="many_to_one")
