#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
01_mise_en_forme_donnees.py

Objectif
--------
Transformer un export Varioskan (Excel) en tableau long avec les colonnes :
    - temps_h
    - souche
    - Groupe
    - replicat
    - DO_brute
    - Lum_brute
    - type   (souche ou blanc)

Le script :
    1. détecte automatiquement la feuille d'absorbance (DO)
    2. utilise, par défaut, la première feuille de luminescence
    3. permet de choisir une autre feuille de luminescence via --lum-sheet
    4. récupère la colonne Groupe depuis la feuille 'Plan de plaque'
    5. reconstruit un tableau long sans faire de moyenne
    6. exporte uniquement un CSV

Exemple d'utilisation
---------------------
python 01_mise_en_forme_donnees.py "Manip_Optimisation_1.xlsx"
python 01_mise_en_forme_donnees.py "Manip_Optimisation_1.xlsx" --lum-sheet "Luminescence 200_04"
python 01_mise_en_forme_donnees.py "Manip_Optimisation_1.xlsx" --output-csv "table_longue.csv"
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook


HEADER_TIME = "temps moy. [s]"
HEADER_LECTURE = "Lecture en cours"
NOMS_LIGNES_PLAQUE = set("ABCDEFGH")


def nettoyer_texte(valeur: object) -> str:
    """Nettoie les espaces et les caractères invisibles."""
    if valeur is None:
        return ""
    texte = str(valeur).replace("\xa0", " ")
    texte = re.sub(r"\s+", " ", texte).strip()
    return texte


def trouver_feuille_absorbance(noms_feuilles: List[str]) -> str:
    candidats = [nom for nom in noms_feuilles if nettoyer_texte(nom).lower().startswith("absorbance")]
    if not candidats:
        raise ValueError("Aucune feuille d'absorbance détectée dans le fichier.")
    return candidats[0]


def trouver_feuilles_luminescence(noms_feuilles: List[str]) -> List[str]:
    return [nom for nom in noms_feuilles if nettoyer_texte(nom).lower().startswith("luminescence")]


def trouver_ligne_entete(ws) -> int:
    """Cherche la ligne qui contient 'Lecture en cours' et 'temps moy. [s]'."""
    for row in range(1, min(ws.max_row, 30) + 1):
        valeurs = [nettoyer_texte(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 10) + 1)]
        if HEADER_LECTURE in valeurs and HEADER_TIME in valeurs:
            return row
    raise ValueError(f"Impossible de trouver la ligne d'entête dans la feuille '{ws.title}'.")


def extraire_table_wide(ws) -> Tuple[pd.DataFrame, List[str]]:
    """Extrait une feuille Varioskan en tableau large."""
    ligne_entete = trouver_ligne_entete(ws)
    entetes = [nettoyer_texte(ws.cell(ligne_entete, col).value) for col in range(1, ws.max_column + 1)]

    colonnes_utiles = [i for i, nom in enumerate(entetes, start=1) if nom != ""]
    if not colonnes_utiles:
        raise ValueError(f"Aucune colonne exploitable détectée dans '{ws.title}'.")

    noms_colonnes = [entetes[i - 1] for i in colonnes_utiles]
    lignes = []

    for row in range(ligne_entete + 1, ws.max_row + 1):
        valeurs = [ws.cell(row, col).value for col in colonnes_utiles]
        if all(v is None for v in valeurs):
            continue

        lecture = valeurs[0] if len(valeurs) > 0 else None
        temps_sec = valeurs[1] if len(valeurs) > 1 else None

        if lecture is None or temps_sec is None:
            continue

        lignes.append(valeurs)

    if not lignes:
        raise ValueError(f"Aucune donnée détectée dans '{ws.title}'.")

    df = pd.DataFrame(lignes, columns=noms_colonnes)
    df = df.rename(columns={HEADER_LECTURE: "lecture", HEADER_TIME: "temps_sec"})

    df["lecture"] = pd.to_numeric(df["lecture"], errors="coerce")
    df["temps_sec"] = pd.to_numeric(df["temps_sec"], errors="coerce")
    df = df.dropna(subset=["lecture", "temps_sec"]).copy()
    df["lecture"] = df["lecture"].astype(int)

    for col in df.columns:
        if col not in {"lecture", "temps_sec"}:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    colonnes_mesures = [c for c in df.columns if c not in {"lecture", "temps_sec"}]
    return df, colonnes_mesures


def parser_echantillon(entete: str) -> Dict[str, str]:
    """Extrait le nom d'échantillon et le puits depuis un en-tête du type 'Nom (B02)'."""
    propre = nettoyer_texte(entete)
    match = re.match(r"^(.*?)(?:\s*\(([A-H]\d{2})\))?$", propre)
    if match:
        nom = nettoyer_texte(match.group(1))
        puits = match.group(2) or ""
    else:
        nom = propre
        puits = ""

    nom_min = nom.lower()
    type_echantillon = "blanc" if ("blanc" in nom_min or "blank" in nom_min) else "souche"

    return {
        "sample_header": entete,
        "souche": nom,
        "puits": puits,
        "type": type_echantillon,
    }


def extraire_groupes_plan_plaque(ws_plan) -> Dict[str, str]:
    """Construit un mapping puits -> Groupe depuis la feuille 'Plan de plaque'."""
    mapping: Dict[str, str] = {}

    for row in range(1, ws_plan.max_row + 1):
        premiere_cellule = nettoyer_texte(ws_plan.cell(row, 1).value)
        if premiere_cellule not in NOMS_LIGNES_PLAQUE:
            continue

        ligne_lettre = premiere_cellule
        row_groupes = row + 1
        if row_groupes > ws_plan.max_row:
            continue

        for col in range(2, min(ws_plan.max_column, 13) + 1):
            numero_colonne = ws_plan.cell(4, col).value
            groupe = nettoyer_texte(ws_plan.cell(row_groupes, col).value)

            if numero_colonne is None or groupe == "":
                continue

            try:
                numero_colonne_int = int(numero_colonne)
            except (TypeError, ValueError):
                continue

            puits = f"{ligne_lettre}{numero_colonne_int:02d}"
            mapping[puits] = groupe

    return mapping



def construire_metadata_echantillons(colonnes_communes: List[str], groupes_par_puits: Dict[str, str]) -> pd.DataFrame:
    meta = pd.DataFrame([parser_echantillon(col) for col in colonnes_communes])
    meta["Groupe"] = meta["puits"].map(groupes_par_puits).fillna("")
    meta["replicat"] = meta.groupby("souche").cumcount() + 1
    return meta



def verifier_ecart_temps(long_df: pd.DataFrame) -> None:
    if "temps_sec_do" not in long_df.columns or "temps_sec_lum" not in long_df.columns:
        return
    ecart = (long_df["temps_sec_do"] - long_df["temps_sec_lum"]).abs()
    ecart_max = float(np.nanmax(ecart)) if len(ecart) else 0.0
    if ecart_max > 5:
        print(
            f"[AVERTISSEMENT] L'écart maximal entre les temps DO et luminescence est de {ecart_max:.2f} s.",
            file=sys.stderr,
        )



def construire_table_longue(
    df_do: pd.DataFrame,
    df_lum: pd.DataFrame,
    colonnes_communes: List[str],
    groupes_par_puits: Dict[str, str],
) -> pd.DataFrame:
    do_long = df_do[["lecture", "temps_sec"] + colonnes_communes].melt(
        id_vars=["lecture", "temps_sec"],
        value_vars=colonnes_communes,
        var_name="sample_header",
        value_name="DO_brute",
    ).rename(columns={"temps_sec": "temps_sec_do"})

    lum_long = df_lum[["lecture", "temps_sec"] + colonnes_communes].melt(
        id_vars=["lecture", "temps_sec"],
        value_vars=colonnes_communes,
        var_name="sample_header",
        value_name="Lum_brute",
    ).rename(columns={"temps_sec": "temps_sec_lum"})

    long_df = do_long.merge(lum_long, on=["lecture", "sample_header"], how="inner")
    verifier_ecart_temps(long_df)

    long_df["temps_sec"] = long_df[["temps_sec_do", "temps_sec_lum"]].mean(axis=1)
    long_df["temps_h"] = long_df["temps_sec"] / 3600.0

    meta = construire_metadata_echantillons(colonnes_communes, groupes_par_puits)
    long_df = long_df.merge(meta, on="sample_header", how="left")

    long_df = long_df[
        [
            "temps_h",
            "souche",
            "Groupe",
            "replicat",
            "DO_brute",
            "Lum_brute",
            "type",
            "puits",
            "lecture",
            "sample_header",
        ]
    ].sort_values(["type", "souche", "replicat", "temps_h"]).reset_index(drop=True)

    return long_df



def determiner_sortie_csv(fichier_entree: Path, nom_feuille_lum: str, output_csv: str | None) -> Path:
    base = fichier_entree.stem
    suffixe = re.sub(r"[^A-Za-z0-9_-]+", "_", nom_feuille_lum).strip("_")
    return Path(output_csv) if output_csv else fichier_entree.with_name(f"{base}_format_long_{suffixe}.csv")



def main() -> None:
    parser = argparse.ArgumentParser(description="Met les données Varioskan au format long.")
    parser.add_argument("fichier_excel", help="Chemin vers le fichier Excel Varioskan.")
    parser.add_argument(
        "--lum-sheet",
        dest="lum_sheet",
        default=None,
        help="Nom exact de la feuille de luminescence à utiliser. Par défaut : première feuille de luminescence détectée.",
    )
    parser.add_argument("--output-csv", dest="output_csv", default=None, help="Chemin du fichier CSV de sortie.")
    args = parser.parse_args()

    fichier_excel = Path(args.fichier_excel)
    if not fichier_excel.exists():
        raise FileNotFoundError(f"Fichier introuvable : {fichier_excel}")

    wb = load_workbook(fichier_excel, data_only=True)
    noms_feuilles = wb.sheetnames

    feuille_do = trouver_feuille_absorbance(noms_feuilles)
    feuilles_lum = trouver_feuilles_luminescence(noms_feuilles)

    if not feuilles_lum:
        raise ValueError("Aucune feuille de luminescence détectée dans le fichier.")

    feuille_lum = args.lum_sheet if args.lum_sheet else feuilles_lum[0]
    if feuille_lum not in noms_feuilles:
        disponibles = "\n - ".join(feuilles_lum)
        raise ValueError(
            f"Feuille de luminescence introuvable : '{feuille_lum}'\n"
            f"Feuilles disponibles :\n - {disponibles}"
        )

    if "Plan de plaque" not in noms_feuilles:
        raise ValueError("La feuille 'Plan de plaque' est introuvable dans le fichier.")

    df_do, cols_do = extraire_table_wide(wb[feuille_do])
    df_lum, cols_lum = extraire_table_wide(wb[feuille_lum])
    groupes_par_puits = extraire_groupes_plan_plaque(wb["Plan de plaque"])

    colonnes_communes = [col for col in cols_do if col in cols_lum]
    if not colonnes_communes:
        raise ValueError("Aucune colonne d'échantillon commune entre la DO et la luminescence.")

    colonnes_do_seules = sorted(set(cols_do) - set(cols_lum))
    colonnes_lum_seules = sorted(set(cols_lum) - set(cols_do))

    if colonnes_do_seules:
        print(f"[INFO] Colonnes présentes seulement en DO : {len(colonnes_do_seules)}", file=sys.stderr)
    if colonnes_lum_seules:
        print(f"[INFO] Colonnes présentes seulement en luminescence : {len(colonnes_lum_seules)}", file=sys.stderr)

    table_longue = construire_table_longue(df_do, df_lum, colonnes_communes, groupes_par_puits)
    csv_path = determiner_sortie_csv(fichier_excel, feuille_lum, args.output_csv)

    table_longue.to_csv(csv_path, index=False, encoding="utf-8-sig")

    resume = (
        table_longue[["souche", "Groupe", "type", "replicat"]]
        .drop_duplicates()
        .sort_values(["type", "souche", "replicat"])
    )

    print("Fichier traité avec succès.")
    print(f"Feuille DO            : {feuille_do}")
    print(f"Feuille luminescence  : {feuille_lum}")
    print(f"Feuille plan de plaque: Plan de plaque")
    print(f"Sortie CSV            : {csv_path}")
    print(f"Nombre de lignes      : {len(table_longue)}")
    print(f"Temps mesurés         : {table_longue['temps_h'].nunique()}")
    print(f"Échantillons          : {resume['souche'].nunique()}")
    print(f"Puits avec groupe     : {(table_longue['Groupe'] != '').sum()}")
    print("\nRésumé des échantillons détectés :")
    print(resume.to_string(index=False))


if __name__ == "__main__":
    main()
