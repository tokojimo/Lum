#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
00_fusion_plaques_endpoint_en_cinetique.py

Objectif
--------
Fusionner plusieurs exports Varioskan "endpoint" (une plaque 96 puits par temps)
en un faux export cinétique compatible avec 01_mise_en_forme_donnees.py.

Cas typique
-----------
Tu as par exemple :
    260626_SCFM2Po_Rep1_t0.xlsx  -> 0 h
    260626_SCFM2Po_Rep1_t1.xlsx  -> 0.5 h
    ...
    260626_SCFM2Po_Rep1_t8.xlsx  -> 6 h

Le script crée un classeur Excel avec :
    - une feuille Absorbance...
    - une feuille Luminescence...
    - une feuille Plan de plaque

Ces feuilles reprennent la structure attendue par le script 01 :
    - ligne d'entête avec "Lecture en cours" et "temps moy. [s]"
    - colonnes d'échantillons du type "Nom (B01)"
    - une ligne par temps expérimental

Important
---------
Les puits physiques changent entre les temps. Le script crée donc des
"puits virtuels" stables par souche et réplicat, afin que les scripts 01-06
puissent fonctionner sans modification.

Pour la manip 260626, les groupes avaient été définis par temps. Par défaut,
ce script corrige en groupes par souche :
    - une souche = un Groupe
    - les blancs sont associés à tous les Groupes de souches

Usage
-----
python 00_fusion_plaques_endpoint_en_cinetique.py 260626_SCFM2Po_Rep1_t*.xlsx
python 00_fusion_plaques_endpoint_en_cinetique.py 260626_SCFM2Po_Rep1_t*.xlsx --output-excel fusion.xlsx
python 00_fusion_plaques_endpoint_en_cinetique.py 260626_SCFM2Po_Rep1_t*.xlsx --group-mode keep
python 00_fusion_plaques_endpoint_en_cinetique.py 260626_SCFM2Po_Rep1_t*.xlsx --time-map "t0=0,t1=0.5,t2=1,t3=1.5,t4=2,t5=2.5,t6=3,t7=4,t8=6"

Étape suivante
--------------
python 01_mise_en_forme_donnees.py fusion.xlsx
"""

from __future__ import annotations

import argparse
import csv
import glob
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_TIME_MAP: Dict[str, float] = {
    "t0": 0.0,
    "t1": 1.0,
    "t2": 2.0,
    "t3": 3.0,
    "t4": 4.0,
    "t5": 5.0,
    "t6": 6.0,
    "t7": 7.0,
}

ROWS_96 = list("ABCDEFGH")
COLS_96 = list(range(1, 13))
HEADER_LECTURE = "Lecture en cours"
HEADER_TIME = "temps moy. [s]"


@dataclass(frozen=True)
class PlateItem:
    file: Path
    time_key: str
    time_h: float
    physical_well: str
    sample_name: str
    original_group: str
    type_sample: str
    replicate_in_time: int
    virtual_well: str
    group_corrected: str
    do_value: float | int | None
    lum_value: float | int | None


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def natural_key(path: Path) -> Tuple:
    """Tri naturel : ...t2 avant ...t10."""
    parts = re.split(r"(\d+)", path.name)
    return tuple(int(p) if p.isdigit() else p.lower() for p in parts)


def parse_time_map(text: str | None) -> Dict[str, float]:
    if not text:
        return dict(DEFAULT_TIME_MAP)
    out: Dict[str, float] = {}
    for chunk in text.split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "=" not in chunk:
            raise ValueError(f"Entrée invalide dans --time-map : {chunk!r}. Format attendu : t0=0")
        key, value = chunk.split("=", 1)
        key = key.strip().lower()
        out[key] = float(value.strip())
    return out


def infer_time_key(path: Path) -> str:
    match = re.search(r"(?:^|[_\-])t(\d+)(?:\D|$)", path.stem, flags=re.IGNORECASE)
    if not match:
        raise ValueError(
            f"Impossible d'inférer le temps depuis le nom du fichier : {path.name}. "
            "Le nom doit contenir t0, t1, t2... ou utiliser un renommage adapté."
        )
    return f"t{int(match.group(1))}"


def find_sheet_name(sheetnames: Iterable[str], prefix: str) -> str:
    prefix = prefix.lower()
    candidates = [s for s in sheetnames if clean_text(s).lower().startswith(prefix)]
    if not candidates:
        raise ValueError(f"Aucune feuille commençant par {prefix!r} détectée.")
    return candidates[0]


def cell_to_float(value: object) -> float | int | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def find_plate_grid(ws) -> Dict[str, float | int | None]:
    """Lit une grille plaque 96 dans une feuille endpoint Varioskan."""
    header_row = None
    for row in range(1, min(ws.max_row, 40) + 1):
        first = clean_text(ws.cell(row, 1).value).lower()
        values = [ws.cell(row, col).value for col in range(2, min(ws.max_column, 13) + 1)]
        numeric_headers = []
        for v in values:
            try:
                numeric_headers.append(int(v))
            except (TypeError, ValueError):
                pass
        if first in {"abs", "rlu"} and set(range(1, 13)).issubset(set(numeric_headers)):
            header_row = row
            break
    if header_row is None:
        raise ValueError(f"Impossible de trouver la grille plaque dans la feuille {ws.title!r}.")

    grid: Dict[str, float | int | None] = {}
    for row in range(header_row + 1, min(ws.max_row, header_row + 8) + 1):
        letter = clean_text(ws.cell(row, 1).value).upper()
        if letter not in ROWS_96:
            continue
        for col in range(2, min(ws.max_column, 13) + 1):
            num_raw = ws.cell(header_row, col).value
            try:
                num = int(num_raw)
            except (TypeError, ValueError):
                continue
            if num not in COLS_96:
                continue
            well = f"{letter}{num:02d}"
            grid[well] = cell_to_float(ws.cell(row, col).value)
    return grid


def read_plan(ws_plan) -> list[tuple[str, str, str]]:
    """Retourne (well, sample_name, original_group)."""
    # Ligne contenant 1..12, normalement ligne 4.
    header_row = None
    for row in range(1, min(ws_plan.max_row, 20) + 1):
        nums = []
        for col in range(2, min(ws_plan.max_column, 13) + 1):
            try:
                nums.append(int(ws_plan.cell(row, col).value))
            except (TypeError, ValueError):
                pass
        if set(range(1, 13)).issubset(set(nums)):
            header_row = row
            break
    if header_row is None:
        raise ValueError("Impossible de trouver l'entête 1..12 dans le Plan de plaque.")

    items: list[tuple[str, str, str]] = []
    for row in range(1, ws_plan.max_row + 1):
        row_label = clean_text(ws_plan.cell(row, 1).value).upper()
        if row_label not in ROWS_96:
            continue
        group_row = row + 1
        for col in range(2, min(ws_plan.max_column, 13) + 1):
            try:
                num = int(ws_plan.cell(header_row, col).value)
            except (TypeError, ValueError):
                continue
            sample_name = clean_text(ws_plan.cell(row, col).value)
            original_group = clean_text(ws_plan.cell(group_row, col).value) if group_row <= ws_plan.max_row else ""
            if not sample_name:
                continue
            items.append((f"{row_label}{num:02d}", sample_name, original_group))
    return items


def is_blank(sample_name: str) -> bool:
    low = sample_name.lower()
    return "blanc" in low or "blank" in low


def assign_virtual_wells(sample_names: list[str], max_reps_by_sample: Dict[str, int]) -> Dict[tuple[str, int], str]:
    """Attribue des puits virtuels stables, 1 ligne par échantillon."""
    mapping: Dict[tuple[str, int], str] = {}
    row_index = 0
    for sample in sample_names:
        if row_index >= len(ROWS_96):
            raise ValueError("Plus de 8 échantillons distincts : adapter assign_virtual_wells pour utiliser plusieurs lignes/blocs.")
        letter = ROWS_96[row_index]
        n_rep = max_reps_by_sample[sample]
        if n_rep > 12:
            raise ValueError(f"Plus de 12 réplicats pour {sample!r}, impossible sur une seule ligne virtuelle.")
        for rep in range(1, n_rep + 1):
            mapping[(sample, rep)] = f"{letter}{rep:02d}"
        row_index += 1
    return mapping


def build_group_mapping(sample_names: list[str], mode: str, original_groups_by_sample: Dict[str, list[str]]) -> Dict[str, str]:
    if mode == "keep":
        out = {}
        for sample in sample_names:
            groups = [g for g in original_groups_by_sample.get(sample, []) if g]
            out[sample] = groups[0] if groups else "Sans_groupe"
        return out

    if mode != "by-strain":
        raise ValueError(f"Mode de groupe non supporté : {mode}")

    non_blanks = [s for s in sample_names if not is_blank(s)]
    strain_groups = {sample: f"Groupe {i}" for i, sample in enumerate(non_blanks, start=1)}
    all_groups = "; ".join(strain_groups.values()) or "Groupe 1"

    out: Dict[str, str] = {}
    for sample in sample_names:
        if is_blank(sample):
            out[sample] = all_groups
        else:
            out[sample] = strain_groups[sample]
    return out


def collect_items(paths: list[Path], time_map: Dict[str, float], group_mode: str) -> tuple[list[PlateItem], list[str], Dict[tuple[str, int], str], Dict[str, str]]:
    raw_records = []
    sample_order: list[str] = []
    original_groups_by_sample: Dict[str, list[str]] = defaultdict(list)
    max_reps_by_sample: Dict[str, int] = defaultdict(int)

    for path in sorted(paths, key=natural_key):
        time_key = infer_time_key(path)
        if time_key not in time_map:
            raise ValueError(f"Aucun temps défini pour {time_key} dans --time-map.")
        time_h = time_map[time_key]

        wb = load_workbook(path, data_only=True, read_only=False)
        sheet_do = find_sheet_name(wb.sheetnames, "Absorbance")
        sheet_lum = find_sheet_name(wb.sheetnames, "Luminescence")
        if "Plan de plaque" not in wb.sheetnames:
            raise ValueError(f"Feuille 'Plan de plaque' absente dans {path.name}")

        grid_do = find_plate_grid(wb[sheet_do])
        grid_lum = find_plate_grid(wb[sheet_lum])
        plan = read_plan(wb["Plan de plaque"])

        per_sample_counter: Dict[str, int] = defaultdict(int)
        for physical_well, sample_name, original_group in plan:
            per_sample_counter[sample_name] += 1
            replicate = per_sample_counter[sample_name]
            if sample_name not in sample_order:
                sample_order.append(sample_name)
            original_groups_by_sample[sample_name].append(original_group)
            max_reps_by_sample[sample_name] = max(max_reps_by_sample[sample_name], replicate)
            raw_records.append({
                "file": path,
                "time_key": time_key,
                "time_h": time_h,
                "physical_well": physical_well,
                "sample_name": sample_name,
                "original_group": original_group,
                "type_sample": "blanc" if is_blank(sample_name) else "souche",
                "replicate_in_time": replicate,
                "do_value": grid_do.get(physical_well),
                "lum_value": grid_lum.get(physical_well),
            })

    virtual_wells = assign_virtual_wells(sample_order, max_reps_by_sample)
    group_by_sample = build_group_mapping(sample_order, group_mode, original_groups_by_sample)

    items: list[PlateItem] = []
    for rec in raw_records:
        items.append(PlateItem(
            file=rec["file"],
            time_key=rec["time_key"],
            time_h=float(rec["time_h"]),
            physical_well=rec["physical_well"],
            sample_name=rec["sample_name"],
            original_group=rec["original_group"],
            type_sample=rec["type_sample"],
            replicate_in_time=int(rec["replicate_in_time"]),
            virtual_well=virtual_wells[(rec["sample_name"], int(rec["replicate_in_time"]))],
            group_corrected=group_by_sample[rec["sample_name"]],
            do_value=rec["do_value"],
            lum_value=rec["lum_value"],
        ))
    return items, sample_order, virtual_wells, group_by_sample


def make_sample_headers(sample_order: list[str], virtual_wells: Dict[tuple[str, int], str]) -> list[tuple[str, str, int, str]]:
    """Retourne (header, sample, rep, virtual_well) dans l'ordre des colonnes."""
    rows = []
    for sample in sample_order:
        reps = sorted(rep for (s, rep), _well in virtual_wells.items() if s == sample)
        for rep in reps:
            well = virtual_wells[(sample, rep)]
            rows.append((f"{sample} ({well})", sample, rep, well))
    return rows


def write_measure_sheet(ws, title: str, headers: list[tuple[str, str, int, str]], times: list[tuple[str, float]], values: Dict[tuple[str, int, str], float | int | None]) -> None:
    ws.title = title
    ws["A1"] = "Résultats de mesure"
    ws["A2"] = "Fusion endpoint -> cinétique virtuelle"
    ws["A5"] = title.split("_")[0]
    ws["A8"] = "Plaque virtuelle"

    header_row = 10
    ws.cell(header_row, 1).value = HEADER_LECTURE
    ws.cell(header_row, 2).value = HEADER_TIME
    for col, (header, _sample, _rep, _well) in enumerate(headers, start=3):
        ws.cell(header_row, col).value = header

    for i, (time_key, time_h) in enumerate(times, start=1):
        row = header_row + i
        ws.cell(row, 1).value = i
        ws.cell(row, 2).value = float(time_h) * 3600.0
        for col, (_header, sample, rep, _well) in enumerate(headers, start=3):
            ws.cell(row, col).value = values.get((time_key, sample, rep))

    style_header(ws, header_row, max_col=len(headers) + 2)


def style_header(ws, header_row: int, max_col: int) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for col in range(1, max_col + 1):
        cell = ws.cell(header_row, col)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = ws.cell(header_row + 1, 3)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    for col in range(3, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 28


def write_plan_sheet(ws, headers: list[tuple[str, str, int, str]], group_by_sample: Dict[str, str]) -> None:
    ws.title = "Plan de plaque"
    ws["A1"] = "Nom"
    ws["B1"] = "Plaque virtuelle"
    ws["A2"] = "Modèle de plaque"
    ws["B2"] = "ANSI/SBS Standard, 96-well"

    for col, num in enumerate(COLS_96, start=2):
        ws.cell(4, col).value = num
        ws.cell(4, col).font = Font(bold=True)

    # Même structure que l'export Varioskan : ligne de noms, ligne suivante de groupes.
    row_for_letter = {
        "A": 5,
        "B": 8,
        "C": 11,
        "D": 14,
        "E": 17,
        "F": 20,
        "G": 23,
        "H": 26,
    }
    for letter, row in row_for_letter.items():
        ws.cell(row, 1).value = letter
        ws.cell(row, 1).font = Font(bold=True)

    for _header, sample, _rep, well in headers:
        letter = well[0]
        col_num = int(well[1:])
        row = row_for_letter[letter]
        col = col_num + 1
        ws.cell(row, col).value = sample
        ws.cell(row + 1, col).value = group_by_sample.get(sample, "")

    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 20 if col > 1 else 14


def write_mapping_csv(path: Path, items: list[PlateItem], headers: list[tuple[str, str, int, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "time_key", "temps_h", "source_file", "sample_name", "type", "replicate",
            "physical_well", "virtual_well", "original_group", "group_corrected", "DO", "Lum"
        ])
        for item in sorted(items, key=lambda x: (x.time_h, x.sample_name, x.replicate_in_time)):
            writer.writerow([
                item.time_key, item.time_h, item.file.name, item.sample_name, item.type_sample,
                item.replicate_in_time, item.physical_well, item.virtual_well,
                item.original_group, item.group_corrected, item.do_value, item.lum_value,
            ])


def determine_output(paths: list[Path], output_excel: str | None) -> Path:
    if output_excel:
        return Path(output_excel)
    first = sorted(paths, key=natural_key)[0]
    stem = re.sub(r"[_\-]?t\d+.*$", "", first.stem, flags=re.IGNORECASE)
    if not stem:
        stem = "fusion_plaques_endpoint"
    return first.with_name(f"{stem}_fusion_endpoint_cinetique.xlsx")


def expand_input_files(arguments: list[str]) -> list[Path]:
    """Accepte les jokers de type *.xlsx, y compris quand PowerShell les passe tels quels."""
    paths: list[Path] = []
    missing: list[str] = []

    for arg in arguments:
        has_wildcard = any(char in arg for char in "*?[")
        if has_wildcard:
            matches = [Path(p) for p in glob.glob(arg)]
            if not matches:
                missing.append(arg)
            else:
                paths.extend(matches)
        else:
            p = Path(arg)
            if p.exists():
                paths.append(p)
            else:
                missing.append(arg)

    # Déduplication en conservant l'ordre naturel.
    unique = {str(p.resolve()).lower(): p for p in paths}
    paths = sorted(unique.values(), key=natural_key)

    if missing:
        raise FileNotFoundError("Fichiers introuvables : " + ", ".join(missing))
    if not paths:
        raise FileNotFoundError("Aucun fichier Excel endpoint trouvé.")

    return paths


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fusionne des plaques endpoint en fichier Excel cinétique compatible avec 01_mise_en_forme_donnees.py."
    )
    parser.add_argument("fichiers", nargs="+", help="Exports Excel endpoint, par exemple *_t0.xlsx *_t1.xlsx ...")
    parser.add_argument("--output-excel", default=None, help="Chemin du fichier Excel fusionné à créer.")
    parser.add_argument(
        "--time-map",
        default=None,
        help="Mapping temps, ex. 't0=0,t1=0.5,t2=1,t3=1.5,t4=2,t5=2.5,t6=3,t7=4,t8=6'."
    )
    parser.add_argument(
        "--group-mode",
        choices=["by-strain", "keep"],
        default="by-strain",
        help="by-strain corrige les groupes par souche; keep conserve le premier groupe observé par échantillon. Défaut : by-strain."
    )
    parser.add_argument(
        "--mapping-csv",
        default=None,
        help="Chemin du CSV de traçabilité. Défaut : même nom que l'Excel avec suffixe _mapping.csv."
    )
    args = parser.parse_args()

    paths = expand_input_files(args.fichiers)

    time_map = parse_time_map(args.time_map)
    items, sample_order, virtual_wells, group_by_sample = collect_items(paths, time_map, args.group_mode)
    headers = make_sample_headers(sample_order, virtual_wells)

    times = sorted({(item.time_key, item.time_h) for item in items}, key=lambda x: x[1])
    do_values = {(item.time_key, item.sample_name, item.replicate_in_time): item.do_value for item in items}
    lum_values = {(item.time_key, item.sample_name, item.replicate_in_time): item.lum_value for item in items}

    output_path = determine_output(paths, args.output_excel).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb_out = Workbook()
    ws_do = wb_out.active
    write_measure_sheet(ws_do, "Absorbance 2_01", headers, times, do_values)
    ws_lum = wb_out.create_sheet("Luminescence 1000_02")
    write_measure_sheet(ws_lum, "Luminescence 1000_02", headers, times, lum_values)
    ws_plan = wb_out.create_sheet("Plan de plaque")
    write_plan_sheet(ws_plan, headers, group_by_sample)

    ws_info = wb_out.create_sheet("Fusion_info")
    ws_info["A1"] = "Fichier source"
    ws_info["B1"] = "time_key"
    ws_info["C1"] = "temps_h"
    for col in range(1, 4):
        ws_info.cell(1, col).font = Font(bold=True)
    for r, (time_key, time_h) in enumerate(times, start=2):
        source_names = sorted({item.file.name for item in items if item.time_key == time_key})
        ws_info.cell(r, 1).value = "; ".join(source_names)
        ws_info.cell(r, 2).value = time_key
        ws_info.cell(r, 3).value = time_h
    ws_info["E1"] = "Échantillon"
    ws_info["F1"] = "Groupe corrigé"
    ws_info["G1"] = "Réplicats"
    for col in range(5, 8):
        ws_info.cell(1, col).font = Font(bold=True)
    for r, sample in enumerate(sample_order, start=2):
        reps = [rep for (s, rep) in virtual_wells if s == sample]
        ws_info.cell(r, 5).value = sample
        ws_info.cell(r, 6).value = group_by_sample[sample]
        ws_info.cell(r, 7).value = len(reps)
    for col in range(1, 8):
        ws_info.column_dimensions[get_column_letter(col)].width = 32

    wb_out.save(output_path)

    mapping_path = Path(args.mapping_csv).resolve() if args.mapping_csv else output_path.with_name(output_path.stem + "_mapping.csv")
    write_mapping_csv(mapping_path, items, headers)

    print("[OK] Fusion endpoint -> cinétique terminée.")
    print(f"[OK] Excel fusionné : {output_path}")
    print(f"[OK] Mapping CSV    : {mapping_path}")
    print(f"[INFO] Temps        : {', '.join([f'{k}={v:g}h' for k, v in times])}")
    print(f"[INFO] Échantillons : {len(sample_order)}")
    print(f"[INFO] Colonnes     : {len(headers)}")
    print("[INFO] Groupes corrigés :")
    for sample in sample_order:
        print(f"  - {sample} -> {group_by_sample[sample]}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"[ERREUR] {exc}", file=sys.stderr)
        raise
